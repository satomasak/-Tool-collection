#!/usr/bin/env python
# coding: utf-8

from selenium import webdriver
import time

#自動操作開始　ページにアクセスし、ログインする

browser = webdriver.Chrome(executable_path = 'webdriverのパス')
browser.implicitly_wait(3)
url = "アクセスするサイトのURL"
browser.get(url)
time.sleep(3)

#ユーザー名とパスワードを入力しクリック（name属性がある場合）
log_name = browser.find_element_by_name('name')
log_name.clear()
log_name.send_keys('ユーザー名')

log_pass = browser.find_element_by_name('password')
log_pass.clear()
log_pass.send_keys('パスワード')

btn_login = browser.find_element_by_xpath('xpathの貼り付け')
time.sleep(1)

btn_login.click()
time.sleep(3)


#トップページに移動
top_page = browser.find_element_by_xpath('xpathの貼り付け')
top_page.click()
time.sleep(3)

#出荷管理のページへ移動
syukka = browser.find_element_by_xpath('xpathの貼り付け')
syukka.click()
time.sleep(3)
syukka_kanri = browser.find_element_by_xpath('xpathの貼り付け')
syukka_kanri.click()
time.sleep(3)


import openpyxl as xl
import datetime


#エクセルの操作　エクセルの表の値をwebのシステムに入力保存する。

#data_only=True:関数の結果を値として取得する(記述しないと関数式が読み込まれる)
wb = xl.load_workbook('エクセルファイルのパス', data_only=True)
st = wb['Sheet1']

#表の入力処理。空白セルでは処理をpassする。

for c in range(5,12):
    for r in range(4,14):

        #納品日の変数代入（print用）
        date_end = st.cell(row=15, column=c).value
        date_end = date_end.strftime('%Y-%m-%d')
        #空白判定
        if st.cell(row=r, column=c).value == None:
            pass
        else:
        #レコードの新規作成
            create_rec = browser.find_element_by_link_text('aタグで囲まれたテキスト')
            time.sleep(2)
            create_rec.click()

            #出荷日の入力
            date_strt = st.cell(row=14, column=c).value
            date_strt = date_strt.strftime('%Y-%m-%d')
            date_out = browser.find_element_by_xpath('xpathの貼り付け')
            date_out.clear()
            date_out.send_keys(date_strt)

            #納品日の入力
            date_in = browser.find_element_by_xpath('xpathの貼り付け')
            date_in.clear()
            date_in.send_keys(date_end)

            #取引先コード入力
            conp_num = browser.find_element_by_id('31_2560-:px-text')
            conp_num.send_keys('1')
            conp_btn = browser.find_element_by_xpath('xpathの貼り付け')
            conp_btn.click()

            #商品コードの入力
            prod_num = browser.find_element_by_xpath('xpathの貼り付け')
            prod_num.clear()
            num = st.cell(row=r, column=4).value
            prod_num.send_keys(num)
            prod_btn = browser.find_element_by_xpath('xpathの貼り付け')
            prod_btn.click()

            #荷姿の異なる商品の処理分岐
            if r == 4:
                #16入り処理
                #数量入力
                prod_quan = browser.find_element_by_xpath('xpathの貼り付け')
                prod_quan.clear()
                valu = st.cell(row=r, column=c).value
                valu = valu*16
                prod_quan.send_keys(valu)
                #セーブ
                save = browser.find_element_by_xpath('xpathの貼り付け')
                save.click()
                time.sleep(3)
            else:
		#20入り処理
                prod_quan = browser.find_element_by_xpath('xpathの貼り付け')
                prod_quan.clear()
                valu = st.cell(row=r, column=c).value
                valu = valu*20
                prod_quan.send_keys(valu)
                #セーブ
                save = browser.find_element_by_xpath('xpathの貼り付け')
                save.click()
                time.sleep(3)
	    time.sleep(3)
	    back_page = browser.find_element_by_link_text('aタグで囲まれたテキスト')
            back_page.click()
            time.sleep(2)
    print(date_end+'納品のデータ入力完了')
print('データ入力すべて終了')





