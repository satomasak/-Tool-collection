from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.common.alert import Alert

#グーグルクロームのwebドライバーのパス
browser = webdriver.Chrome(executable_path = 'ドライバーの絶対パス')
browser.implicitly_wait(5)
#待機の最大時間を設定　10秒
#読み込みが完了すれば10秒以内で順次処理させるためtime.sleepを使わない
wait = WebDriverWait(browser, 10)
#指定サイトへの移動
url = "サイトURL"
browser.get(url)

#ページが読み込まれるまで待機
wait.until(ec.presence_of_all_elements_located)
#ログイン
log_name = browser.find_element_by_name('username')
log_name.clear()
log_name.send_keys('ログイン名i')
log_pass = browser.find_element_by_name('password')
log_pass.clear()
log_pass.send_keys('パスワード')
btn_login = browser.find_element_by_xpath('xパス')
btn_login.click()

#ページが読み込まれるまで待機
wait.until(ec.presence_of_all_elements_located)
#トップページへの移動
top_page = browser.find_element_by_xpath('xパス')
top_page.click()

#ページが読み込まれるまで待機
wait.until(ec.presence_of_all_elements_located)
syukka = browser.find_element_by_xpath('xパス')
syukka.click()

#ページが読み込まれるまで待機
wait.until(ec.presence_of_all_elements_located)
syukka_kanri = browser.find_element_by_xpath('xパス')
syukka_kanri.click()

#ページが読み込まれるまで待機
wait.until(ec.presence_of_all_elements_located)
import openpyxl as xl
import datetime
#data_only=True:関数の結果を値として取得する(記述しないと関数式が読み込まれる)
wb = xl.load_workbook('入力用データの書かれたエクセルの絶対パス', data_only=True)
st = wb['Sheet1']
import time
#表の処理。10×７
for c in range(6,13):
    for r in range(4,14):
        #納品日の変数代入（print用）
        date_end = st.cell(row=15, column=c).value
        date_end = date_end.strftime('%Y-%m-%d')
        #空白判定
        if st.cell(row=r, column=c).value == None:
            pass
        else:
        #レコードの新規作成
            create_rec = browser.find_element_by_link_text('aリンクのテキスト')
            create_rec.click()

            #ページが読み込まれるまで待機
            wait.until(ec.presence_of_all_elements_located)

            #出荷日の入力
            date_strt = st.cell(row=14, column=c).value
            date_strt = date_strt.strftime('%Y-%m-%d')
            date_out = browser.find_element_by_xpath('xパス')
            date_out.clear()
            date_out.send_keys(date_strt)

            #納品日の入力
            date_in = browser.find_element_by_xpath('xパス')
            date_in.clear()
            date_in.send_keys(date_end)

            #取引先コード入力
            conp_num = browser.find_element_by_id('id名')
            conp_num.send_keys('取引先コード')
            conp_btn = browser.find_element_by_xpath('xパス')
            conp_btn.click()

            #商品コードの入力
            prod_num = browser.find_element_by_xpath('xパス')
            prod_num.clear()
            num = st.cell(row=r, column=4).value
            prod_num.send_keys(num)
            prod_btn = browser.find_element_by_xpath('xパス')
            prod_btn.click()

            #単価の入力
            price = browser.find_element_by_xpath('xパス')
            price.clear()
            price_num = st.cell(row=r, column=5).value
            price.send_keys(price_num)

            #荷姿の判定
            if r == 4:
                #16入り処理
                #数量入力
                prod_quan = browser.find_element_by_xpath('xパス')
                prod_quan.clear()
                valu = st.cell(row=r, column=c).value
                valu = valu*16
                prod_quan.send_keys(valu)

            else:
            	#20入り商品
                prod_quan = browser.find_element_by_xpath('xパス')
                prod_quan.clear()
                valu = st.cell(row=r, column=c).value
                valu = valu*20
                prod_quan.send_keys(valu)

            #ポップアップアラート処理のための例外処理
            try:
                 #セーブの段階でアラートが表示されることがあるため例外処理を行う
                save = browser.find_element_by_xpath('xパス')
                save.click()
                #ページが読み込まれるまで待機
                time.sleep(2)
                Alert(browser).accept()
                print("alert accepted")
                #ここの待機はtimesleepでないと要素が見つからないというエラーが出る　なぜかわからない
                time.sleep(2)
                back_page = browser.find_element_by_link_text('aリンクのテキスト')
                back_page.click()
                #ページが読み込まれるまで待機
                wait.until(ec.presence_of_all_elements_located)
            except NoAlertPresentException:
                 #ここの待機はtimesleepでないと要素が見つからないというエラーが出る　なぜかわからない
                time.sleep(2)
                back_page = browser.find_element_by_link_text('aリンクのテキスト')
                back_page.click()
                #ページが読み込まれるまで待機
                wait.until(ec.presence_of_all_elements_located)
    print(date_end+'納品のデータ入力完了')
print('データ入力すべて終了')
