#エクセルのロードとセーブ
import openpyxl as xl

#data_only=True:関数の結果を値として取得する(記述しないと関数式が読み込まれる)
wb = xl.load_workbook('ファイルパス', data_only=True)
st = wb['Sheet1']


wb.save('ファイルパス')



#pandasを使ったロードとセーブ
import pandas as pd

selldf = pd.read_excel('ファイルパス',usecols=[1,5,6,9,11,12,15,17,18])



with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
	df名.to_excel(writer, sheet_name=i)



#エクセルを新しく作成
import openpyxl as xl

wb = xl.Workbook()
wb.create_sheet('シート名')

wb.save('任意のファイルパス+任意のファイル名')